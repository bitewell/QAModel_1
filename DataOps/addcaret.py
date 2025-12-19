
# Imports

import openpyxl
import csv
import os
from pathlib import Path



# Function: Apply Caret Modifications

def apply_caret_modifications(input_xlsx):

    # Print the file name being processed
    print(f"Reading tab 'all_scores' from {input_xlsx}...")
    
    # Load the workbook
    try:
        wb = openpyxl.load_workbook(input_xlsx, data_only=True)
    except FileNotFoundError:
        print("Error: The .xlsx file was not found.")
        return None, 0

    
    # Get the target sheet
    target_sheet = "all_scores"
    if target_sheet not in wb.sheetnames:
        print(f"Error: Tab '{target_sheet}' not found. Available tabs: {wb.sheetnames}")
        return None, 0
    sheet = wb[target_sheet]


    # Define target headers to check for color formatting
    target_headers = [
        "category_name", "product_name", "ingredients_text", "calories", "total_fat",

        "sat_fat", "trans_fat", "unsat_fat", "cholesterol", "sodium", "carbs",

        "dietary_fiber", "total_sugars", "added_sugars", "protein", "potassium",

        "is_whole_grain", "is_omega_three", "is_healthy_oils", "is_healthy_fats",

        "is_seed_oil", "is_refined_grains", "is_deep_fried", "is_sugars_added",

        "is_artificial_sweeteners", "is_artificial_flavors", "is_artificial_preservatives",

        "is_artificial_colors", "is_artificial_red_color", "is_ph_oil", "is_aspartame",

        "is_acesulfame_potassium", "is_saccharin", "is_corn_syrup", "is_brominated_vegetable_oil",

        "is_potassium_bromate", "is_titanium_dioxide", "is_phosphate_additives",

        "is_polysorbate60", "is_mercury_fish", "is_caregeenan", "is_natural_non_kcal_sweeteners",

        "is_natural_additives", "is_unspecific_ingredient", "is_propellant", "is_starch",

        "is_active_live_cultures"
    ]

    # Map headers to column indices in the sheet
    header_row = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
    target_indices = [i for i, h in enumerate(header_row) if h in target_headers]

    # Initialize variables for storing processed data
    csv_data = []
    caret_count = 0

    
    # Iterate through all rows and columns
    for r_idx, row in enumerate(sheet.iter_rows(values_only=False), start=1):
        row_values = []

        for c_idx, cell in enumerate(row):
            # Get cell value, default to empty string if None
            val = cell.value if cell.value is not None else ""

            # Convert float values to integers if they're whole numbers
            # This removes the .0 suffix from numbers like 5.0 -> "5"
            if isinstance(val, float) and val.is_integer():
                val = str(int(val))
            else:
                val = str(val) if val != "" else ""

            
            # Apply caret logic for target columns (skip header row)
            if r_idx > 1 and c_idx in target_indices:
                # Get cell fill color
                fill = cell.fill
                color = None

                if fill and fill.start_color:
                    # Get the RGB value (openpyxl stores as ARGB, e.g., 'FFFFFF00' for yellow)
                    color = fill.start_color.rgb if fill.start_color.type == 'rgb' else None

                # Target colors: yellow (#FFFF00) and cyan (#00FFFF)
                target_colors = ['FFFF00', '00FFFF']
                is_target_color = color and str(color)[-6:].upper() in target_colors

                # If the cell has a yellow or cyan background, add a caret (^) to the value
                if is_target_color:
                    val_stripped = val.strip()
                    # Only add caret if value exists and doesn't already end with caret
                    if val_stripped and not val_stripped.endswith('^'):
                        val = f"{val_stripped}^"
                        caret_count += 1

            # Add the processed value to the row
            row_values.append(val)

        # Add the completed row to the CSV data
        csv_data.append(row_values)

    print(f"Success! Added {caret_count} carets.")
    return csv_data, caret_count



# Function: Save to CSV
def save_to_csv(csv_data, output_csv):

    # Check if data exists
    if csv_data is None:
        print("Error: No data to save.")
        return

    # Write the data to CSV file
    with open(output_csv, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(csv_data)

    print(f"File saved as: {output_csv}")



# Function: Process Scores to CSV with Caret
def process_scores_to_csv_with_caret(input_xlsx, output_csv):
    csv_data, caret_count = apply_caret_modifications(input_xlsx)
    save_to_csv(csv_data, output_csv)




if __name__ == "__main__":
    # Define folder paths
    preop_data_folder = Path(__file__).parent.parent / "Data" / "PreOpData"
    output_folder = Path(__file__).parent.parent / "Data" / "PreOpDataCSV"
    
    # Create output folder if it doesn't exist
    output_folder.mkdir(parents=True, exist_ok=True)
    
    # Find all After*.xlsx and Before*.xlsx files
    after_files = sorted(preop_data_folder.glob("After*.xlsx"))
    before_files = sorted(preop_data_folder.glob("Before*.xlsx"))
    
    # Combine all files to process
    all_files = after_files + before_files
    
    # Check if any files were found
    if not all_files:
        print(f"No After*.xlsx or Before*.xlsx files found in {preop_data_folder}")
    else:
        # Display found files
        print(f"Found {len(after_files)} After files and {len(before_files)} Before files to process:")
        for file in all_files:
            print(f"  - {file.name}")
        print()
        
        # Process each file
        for input_file in all_files:
            # Generate output CSV filename (e.g., After1.xlsx -> After1.csv)
            output_filename = input_file.stem + ".csv"
            output_path = output_folder / output_filename
            
            # Display processing status
            print(f"\nProcessing: {input_file.name}")
            
            # Process and save the file
            process_scores_to_csv_with_caret(str(input_file), str(output_path))
            print()
