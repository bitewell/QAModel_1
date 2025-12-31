import openpyxl
import csv
from pathlib import Path


def apply_caret_modifications(input_xlsx):

    # Reads Excel file and adds caret (^) markers to cells with yellow or cyan background colors
    print(f"Reading tab 'all_scores' from {input_xlsx}...")
    
    # Load the workbook
    try:
        wb = openpyxl.load_workbook(input_xlsx, data_only=True)
    except FileNotFoundError:
        print("Error: The .xlsx file was not found.")
        return None, 0

    target_sheet = "all_scores"
    if target_sheet not in wb.sheetnames:
        print(f"Error: Tab '{target_sheet}' not found. Available tabs: {wb.sheetnames}")
        return None, 0


    sheet = wb[target_sheet]


    # Columns that can have carets added based on cell colors
    # Only sodium and text columns for sodium-specific anomaly detection
    target_headers = [
        "category_name", "product_name", "ingredients_text", "sodium"
    ]

    # Map headers to column indices
    header_row = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
    target_indices = [i for i, h in enumerate(header_row) if h in target_headers]

    csv_data = []
    caret_count = 0

    # Process all rows and columns
    for r_idx, row in enumerate(sheet.iter_rows(values_only=False), start=1):
        row_values = []

        for c_idx, cell in enumerate(row):
            val = cell.value if cell.value is not None else ""

            # Convert float values to integers if they're whole numbers
            if isinstance(val, float) and val.is_integer():
                val = str(int(val))
            else:
                val = str(val) if val != "" else ""

            # Check for yellow or cyan background colors in target columns (skip header row)
            if r_idx > 1 and c_idx in target_indices:
                fill = cell.fill
                color = None

                if fill and fill.start_color:
                    # openpyxl stores colors as ARGB format
                    color = fill.start_color.rgb if fill.start_color.type == 'rgb' else None

                # Check for yellow (#FFFF00) or cyan (#00FFFF) - check last 6 chars of ARGB
                target_colors = ['FFFF00', '00FFFF']
                is_target_color = color and str(color)[-6:].upper() in target_colors

                # Add caret to values with target background colors
                if is_target_color:
                    val_stripped = val.strip()
                    if val_stripped and not val_stripped.endswith('^'):
                        val = f"{val_stripped}^"
                        caret_count += 1

            row_values.append(val)

        csv_data.append(row_values)

    print(f"Success! Added {caret_count} carets.")
    return csv_data, caret_count


def save_to_csv(csv_data, output_csv):
    # Saves processed data to CSV file
    if csv_data is None:
        print("Error: No data to save.")
        return

    with open(output_csv, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(csv_data)

    print(f"File saved as: {output_csv}")


def process_scores_to_csv_with_caret(input_xlsx, output_csv):
    # Main function: applies caret modifications and saves to CSV
    csv_data, caret_count = apply_caret_modifications(input_xlsx)
    save_to_csv(csv_data, output_csv)


if __name__ == "__main__":
    preop_data_folder = Path(__file__).parent.parent / "DataSodium" / "PreOpData"
    output_folder = Path(__file__).parent.parent / "DataSodium" / "PreOpDataCSV"
    
    output_folder.mkdir(parents=True, exist_ok=True)
    
    # Find all Before and After Excel files
    after_files = sorted(preop_data_folder.glob("After*.xlsx"))
    before_files = sorted(preop_data_folder.glob("Before*.xlsx"))
    all_files = after_files + before_files
    
    if not all_files:
        print(f"No After*.xlsx or Before*.xlsx files found in {preop_data_folder}")
    else:
        print(f"Found {len(after_files)} After files and {len(before_files)} Before files to process:")
        for file in all_files:
            print(f"  - {file.name}")
        print()
        
        # Process each file
        for input_file in all_files:
            output_filename = input_file.stem + ".csv"
            output_path = output_folder / output_filename
            
            print(f"\nProcessing: {input_file.name}")
            process_scores_to_csv_with_caret(str(input_file), str(output_path))
            print()
